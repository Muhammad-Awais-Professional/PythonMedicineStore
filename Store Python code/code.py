import tkinter as tk
from tkinter import ttk
import openpyxl
from tkinter import Toplevel
from tkinter import messagebox

# Dictionary to store medicine information (using unique key)
medicine_dict = {}

def calculate_total_price():
    try:
        name = name_entry.get()
        mg = int(mg_entry.get())
        price_per_box = float(price_per_box_entry.get())
        tablets_per_box = int(tablets_per_box_entry.get())
        tablets_available = int(tablets_available_entry.get())

        price_per_tablet = price_per_box / tablets_per_box
        total_price = price_per_tablet * tablets_available

        result_label.config(text=f"Total Price: ${total_price:.2f}")

        # Check if the medicine already exists in the dictionary
        if (name, mg) in medicine_dict:
            existing_medicine = medicine_dict[(name, mg)]
            existing_medicine["Tablets Available"] += tablets_available
            existing_medicine["Total Price"] += total_price
        else:
            # Create a new entry in the dictionary
            medicine_info = {
                "Name": name,
                "Mg of Medicine": mg,
                "Price per Box": price_per_box,
                "Tablets per Box": tablets_per_box,
                "Tablets Available": tablets_available,
                "Total Price": total_price
            }
            medicine_dict[(name, mg)] = medicine_info

        # Refresh the medicine list in the UI
        refresh_medicine_list()

        # Save the information to Excel
        save_to_excel()

    except ValueError:
        result_label.config(text="Invalid input")

def refresh_medicine_list():
    medicine_listbox.delete(0, tk.END)
    for name, mg in medicine_dict:
        medicine_listbox.insert(tk.END, f"{name} - {mg} mg")

def save_to_excel():
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Name", "Mg of Medicine", "Price per Box", "Tablets per Box", "Tablets Available", "Total Price"])

    for medicine_info in medicine_dict.values():
        ws.append([
            medicine_info["Name"],
            medicine_info["Mg of Medicine"],
            medicine_info["Price per Box"],
            medicine_info["Tablets per Box"],
            medicine_info["Tablets Available"],
            medicine_info["Total Price"]
        ])

    excel_filename = "medicine_info.xlsx"
    wb.save(excel_filename)

def delete_selected_medicine():
    selected_index = medicine_listbox.curselection()
    if selected_index:
        selected_medicine = medicine_listbox.get(selected_index[0])
        name, mg = selected_medicine.split(" - ")
        name = name.strip()
        mg = int(mg.split()[0])
        if (name, mg) in medicine_dict:
            del medicine_dict[(name, mg)]
            refresh_medicine_list()
            save_to_excel()


def load_data_from_excel():
    try:
        excel_filename = "medicine_info.xlsx"
        wb = openpyxl.load_workbook(excel_filename)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, mg, price_per_box, tablets_per_box, tablets_available, total_price = row
            mg = int(mg)
            price_per_box = float(price_per_box)
            tablets_per_box = int(tablets_per_box)
            tablets_available = int(tablets_available)
            total_price = float(total_price)

            medicine_info = {
                "Name": name,
                "Mg of Medicine": mg,
                "Price per Box": price_per_box,
                "Tablets per Box": tablets_per_box,
                "Tablets Available": tablets_available,
                "Total Price": total_price
            }

            medicine_dict[(name, mg)] = medicine_info

    except Exception as e:
        print(f"Error loading data from Excel: {e}")

# Load data from Excel before creating the UI
load_data_from_excel()

def reset_all_medicines():
    # Display a confirmation dialog
    response = messagebox.askyesno("Confirmation", "Are you sure you want to reset all medicines?")
    if response:
        global medicine_dict
        medicine_dict = {}
        refresh_medicine_list()
        save_to_excel()


def edit_selected_medicine():
    # Get the selected medicine
    selected_index = medicine_listbox.curselection()
    if not selected_index:
        return

    selected_medicine = medicine_listbox.get(selected_index[0])
    name, mg = selected_medicine.split(" - ")
    name = name.strip()
    mg = int(mg.split()[0])

    if (name, mg) not in medicine_dict:
        return

    medicine_info = medicine_dict[(name, mg)]

    # Create a new Toplevel window
    edit_window = Toplevel(root)
    edit_window.title(f"Edit {name} - {mg}mg")

    # Fill the window with input fields pre-filled with the medicine's details
    new_name_label = ttk.Label(edit_window, text="Name of the medicine:")
    new_name_label.pack()
    new_name_entry = ttk.Entry(edit_window)
    new_name_entry.insert(0, name)
    new_name_entry.pack()

    new_mg_label = ttk.Label(edit_window, text="Mg of the medicine:")
    new_mg_label.pack()
    new_mg_entry = ttk.Entry(edit_window)
    new_mg_entry.insert(0, mg)
    new_mg_entry.pack()

    new_price_per_box_label = ttk.Label(edit_window, text="Price per box of medicine:")
    new_price_per_box_label.pack()
    new_price_per_box_entry = ttk.Entry(edit_window)
    new_price_per_box_entry.insert(0, medicine_info["Price per Box"])
    new_price_per_box_entry.pack()

    new_tablets_per_box_label = ttk.Label(edit_window, text="Tablets per box:")
    new_tablets_per_box_label.pack()
    new_tablets_per_box_entry = ttk.Entry(edit_window)
    new_tablets_per_box_entry.insert(0, medicine_info["Tablets per Box"])
    new_tablets_per_box_entry.pack()

    new_tablets_available_label = ttk.Label(edit_window, text="Tablets available:")
    new_tablets_available_label.pack()
    new_tablets_available_entry = ttk.Entry(edit_window)
    new_tablets_available_entry.insert(0, medicine_info["Tablets Available"])
    new_tablets_available_entry.pack()

    new_total_price_label = ttk.Label(edit_window, text="Total Price:")
    new_total_price_label.pack()
    total_price_value_label = ttk.Label(edit_window, text=f"${medicine_info['Total Price']:.2f}")
    total_price_value_label.pack()

    def save_changes():
        # Get the edited details
        new_name = new_name_entry.get()
        new_mg = int(new_mg_entry.get())
        new_price_per_box = float(new_price_per_box_entry.get())
        new_tablets_per_box = int(new_tablets_per_box_entry.get())
        new_tablets_available = int(new_tablets_available_entry.get())
        new_total_price = new_price_per_box * new_tablets_available / new_tablets_per_box

        # Update the medicine_dict with the new details
        medicine_dict[(new_name, new_mg)] = {
            "Name": new_name,
            "Mg of Medicine": new_mg,
            "Price per Box": new_price_per_box,
            "Tablets per Box": new_tablets_per_box,
            "Tablets Available": new_tablets_available,
            "Total Price": new_total_price
        }

        # Close the Toplevel window
        edit_window.destroy()

        # Refresh the main window's medicine list
        refresh_medicine_list()

        # Save the updated data to Excel
        save_to_excel()

    save_button = ttk.Button(edit_window, text="Save Changes", command=save_changes)
    save_button.pack()

from tkinter import font

# Create the main window
root = tk.Tk()
root.title("Medicine Records")

# Set a consistent font for headers
header_font = font.Font(size=14, weight='bold')

# Create a frame for medicine entry
entry_frame = ttk.Frame(root, padding="10")
entry_frame.pack(pady=20)

ttk.Label(entry_frame, text="Medicine Details", font=header_font).grid(row=0, column=1)

ttk.Label(entry_frame, text="Name of the medicine:").grid(row=1, column=0, sticky=tk.W, pady=5)
name_entry = ttk.Entry(entry_frame)
name_entry.grid(row=1, column=1, pady=5, padx=5, sticky=tk.EW)

ttk.Label(entry_frame, text="Price per box of medicine:").grid(row=2, column=0, sticky=tk.W, pady=5)
price_per_box_entry = ttk.Entry(entry_frame)
price_per_box_entry.grid(row=2, column=1, pady=5, padx=5, sticky=tk.EW)

ttk.Label(entry_frame, text="Mg of the medicine:").grid(row=3, column=0, sticky=tk.W, pady=5)
mg_entry = ttk.Entry(entry_frame)
mg_entry.grid(row=3, column=1, pady=5, padx=5, sticky=tk.EW)

ttk.Label(entry_frame, text="Tablets per box:").grid(row=4, column=0, sticky=tk.W, pady=5)
tablets_per_box_entry = ttk.Entry(entry_frame)
tablets_per_box_entry.grid(row=4, column=1, pady=5, padx=5, sticky=tk.EW)

ttk.Label(entry_frame, text="Tablets available:").grid(row=5, column=0, sticky=tk.W, pady=5)
tablets_available_entry = ttk.Entry(entry_frame)
tablets_available_entry.grid(row=5, column=1, pady=5, padx=5, sticky=tk.EW)

calculate_button = ttk.Button(entry_frame, text="Calculate Total Price", command=calculate_total_price)
calculate_button.grid(row=6, column=1, pady=20)

result_label = ttk.Label(entry_frame, text="")
result_label.grid(row=7, columnspan=2)

# Create a frame for list and edit controls
list_frame = ttk.Frame(root, padding="10")
list_frame.pack(pady=20, fill=tk.BOTH, expand=True)

edit_button = ttk.Button(list_frame, text="Edit Selected Medicine", command=edit_selected_medicine)
edit_button.grid(row=0, column=0, pady=5, padx=5)

medicine_listbox = tk.Listbox(list_frame)
medicine_listbox.grid(row=1, column=0, columnspan=2, pady=5, padx=5, sticky=tk.EW)

scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=medicine_listbox.yview)
scrollbar.grid(row=1, column=2, sticky=tk.NS)
medicine_listbox.config(yscrollcommand=scrollbar.set)

delete_button = ttk.Button(list_frame, text="Delete Selected Medicine", command=delete_selected_medicine)
delete_button.grid(row=2, column=0, pady=5, padx=5)

reset_button = ttk.Button(list_frame, text="Reset All Medicines", command=reset_all_medicines)
reset_button.grid(row=2, column=1, pady=5, padx=5)

# Start the main loop
root.mainloop()
